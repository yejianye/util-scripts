#!/usr/bin/env python
# Todo
# - [ ] excel intergration, word/translate pairs to excel sheet
# - [ ] youdao translate api

import os
import sys
import hashlib
import uuid
import time
import requests
import argparse

from pprint import pprint

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

CONFIG = {'debug': False}

YOUDAO_URL = 'https://openapi.youdao.com/api'
YOUDAO_APP_ID = "192d9338c6a4d7bc"
YOUDAO_SECRET = "lD0LvnGfgsnwHgvSAqzIhxtO8eo08ufG"
SCRIPT_DIR = '/Users/ryan/utils'
TEMPLATE_FILE = os.path.join(SCRIPT_DIR, "word_translate_template.xlsx")
OUTPUT_PATH = '/Users/ryan/Downloads'

RETRY_COUNT_MAX = 3
RETRY_INTERVAL = 3.0

def encrypt(signStr):
    hash_algorithm = hashlib.sha256()
    hash_algorithm.update(signStr.encode('utf-8'))
    return hash_algorithm.hexdigest()

def truncate(q):
    if q is None:
        return None
    size = len(q)
    return q if size <= 20 else q[0:10] + str(size) + q[size - 10:size]

def cell_style():
    style = NamedStyle(name='default_cell')
    style.font = Font(size=12)
    bd = Side(style='thin', color="000000")
    style.border = Border(top=bd, left=bd, right=bd, bottom=bd)
    style.alignment = Alignment(horizontal="left",
                                vertical="top",
                                wrap_text=True)
    return style

def sign_request(data):
    salt = str(uuid.uuid1())
    curtime = str(int(time.time()))
    signStr = YOUDAO_APP_ID + truncate(data['q']) + salt + curtime + YOUDAO_SECRET
    sign = encrypt(signStr)
    data['appKey'] = YOUDAO_APP_ID
    data['signType'] = 'v3'
    data['curtime'] = curtime
    data['salt'] = salt
    data['sign'] = sign
    return data

def trans_simple(resp):
    if 'translation' in resp:
        return resp['translation']
    else:
        return None

def trans_explains(resp):
    if 'basic' in resp:
        return resp['basic']['explains']
    else:
        return None

def youdao_api_request(data, retry_count=0):
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    resp = requests.post(YOUDAO_URL, data=sign_request(data), headers=headers)
    result = resp.json()
    if result['errorCode'] != '0' and retry_count < RETRY_COUNT_MAX:
        print('Hit rate limit. Pause and retry in {} seconds'.format(RETRY_INTERVAL))
        time.sleep(RETRY_INTERVAL)
        return youdao_api_request(data, retry_count+1)
    return result

def translate_word(word, method=None, debug=False):
    data = {
        'from': 'EN',
        'to': 'zh-CHS',
        'q': word,
    }
    result = youdao_api_request(data)
    if debug:
        pprint(result)

    trans_methods = {
        'simple': trans_simple,
        'explains': trans_explains}
    trans_priority = [
        'simple',
        'explains']

    trans = None
    if method:
        trans = trans_methods[method](result)
    if not trans:
        trans = [trans_methods[m](result) for m in trans_priority]
        trans = [t for t in trans if t is not None]
        trans = trans[0] if trans else ''

    phonetic = result.get('basic', {}).get('phonetic')
    if not phonetic and (' ' in word):
        phonetic = [translate_word(w, debug=debug).get('phonetic') for w in word.split(' ')]
        phonetic = ' '.join(p.split(';')[0] for p in phonetic if p)

    return {
        'translation': trans,
        'phonetic': phonetic}

def set_cell(sheet, row, column, value, style):
    cell = sheet.cell(row=row, column=column, value=value)
    cell.style = style
    return cell

def generate_xlsx(path, title, word_list):
    if os.path.exists(path):
        os.remove(path)
    wb = load_workbook(filename = TEMPLATE_FILE)
    style = cell_style()
    ws = wb.active
    ws.cell(row=1, column=1, value=title)
    for idx, word in enumerate(word_list):
        # word = word.lower()
        trans = translate_word(word, debug=CONFIG['debug'])
        print("Translate {}".format(word))
        set_cell(ws, row=idx+3, column=1, value=word, style=style)
        phonetic, trans = trans.get("phonetic"), trans.get("translation")
        phonetic = '/{}/'.format(phonetic) if phonetic else ''
        set_cell(ws, row=idx+3, column=2, value=phonetic, style=style)
        set_cell(ws, row=idx+3, column=3, value='\n'.join(trans), style=style)
    wb.save(path)

def generate_xlsx_from_stdin():
    lines = sys.stdin.readlines()
    lines = [l.strip() for l in lines if l.strip()]
    title, word_list = lines[0], lines[1:]
    generate_xlsx(os.path.join(OUTPUT_PATH, '{}.xlsx'.format(title)), title, word_list)

def print_translation():
    lines = sys.stdin.readlines()
    word_list = [l.strip() for l in lines if l.strip()]
    for idx, word in enumerate(word_list):
        trans = translate_word(word, debug=CONFIG['debug'])
        phonetic, trans = trans.get("phonetic"), trans.get("translation")
        print("/{}/ {}".format(phonetic or "", ";".join(trans)))

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("--debug", action='store_true', help="Enable debugging")
    parser.add_argument("--stdout", action='store_true', help="Print translations in stdout instead of storing them in Excel")
    args = parser.parse_args()
    CONFIG['debug'] = args.debug
    if args.stdout:
        print_translation()
    else:
        generate_xlsx_from_stdin()
    # translate_word("take into account")
    # generate_xlsx('test.xlsx', ['morning', 'afternoon', 'evening', 'marble'])
